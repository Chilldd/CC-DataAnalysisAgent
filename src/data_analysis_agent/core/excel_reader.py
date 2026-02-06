"""Excel 文件读取器"""

import pandas as pd
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from functools import lru_cache
import time
from collections import OrderedDict

from .logging_config import get_logger, get_metrics

logger = get_logger(__name__)
metrics = get_metrics()


class ExcelReader:
    """Excel 文件读取器，支持读取 Excel 和 CSV 文件"""

    def __init__(
        self,
        file_path: str,
        enable_cache: bool = True,
        max_cache_size: int = 10,
        enable_threading: bool = False
    ):
        """
        初始化 ExcelReader

        Args:
            file_path: Excel 或 CSV 文件路径
            enable_cache: 是否启用缓存（默认启用）
            max_cache_size: 缓存最大条目数（默认10），用于 LRU 清理
            enable_threading: 是否启用多线程读取（默认不启用）
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            logger.error(f"文件不存在: {file_path}")
            raise FileNotFoundError(f"文件不存在: {file_path}")

        self.enable_cache = enable_cache
        self.max_cache_size = max_cache_size
        self.enable_threading = enable_threading

        logger.info(f"初始化 ExcelReader: {file_path}, 缓存={enable_cache}, 多线程={enable_threading}")

        # 使用 OrderedDict 实现 LRU 缓存
        self._cache: OrderedDict[str, tuple[pd.DataFrame, float]] = OrderedDict()  # {key: (df, mtime)}
        self._row_count_cache: OrderedDict[str, tuple[int, float]] = OrderedDict()  # {(sheet_name, usecols): (count, mtime)}

        # 缓存统计
        self._cache_hits: int = 0
        self._cache_misses: int = 0

    def get_info(
        self,
        sheet_name: Optional[str] = None,
        sample_rows: int = 10,
        usecols: Optional[Union[str, List[int], List[str]]] = None
    ) -> Dict[str, Any]:
        """
        获取数据结构信息

        Args:
            sheet_name: 工作表名称（可选，默认第一个）
            sample_rows: 样本行数（默认10）
            usecols: 要读取的列，可以是列名列表、列索引列表或单个列名（可选）

        Returns:
            包含文件信息、列信息、样本数据的字典
        """
        df = self._read_file(sheet_name, usecols=usecols)

        # 获取样本数据并转换为 JSON 可序列化的格式
        sample_df = df.head(sample_rows).copy()
        for col in sample_df.columns:
            if pd.api.types.is_datetime64_any_dtype(sample_df[col]):
                sample_df[col] = sample_df[col].astype(str)

        # 将样本数据转换为 JSON 可序列化的格式
        import numpy as np

        def _to_serializable(val):
            """将 numpy 类型转换为 JSON 可序列化的 Python 原生类型"""
            if pd.isna(val):
                return None
            elif isinstance(val, (np.integer, np.int64, np.int32)):
                return int(val)
            elif isinstance(val, (np.floating, np.float64, np.float32)):
                return float(val)
            elif isinstance(val, (np.bool_, bool)):
                return bool(val)
            elif hasattr(val, 'isoformat'):  # datetime
                return val.isoformat()
            else:
                return val

        # 转换样本数据
        sample_data = []
        for record in sample_df.to_dict('records'):
            sample_data.append({k: _to_serializable(v) for k, v in record.items()})

        return {
            "success": True,
            "file_info": {
                "name": self.file_path.name,
                "size": self.file_path.stat().st_size,
                "sheets": self._get_sheet_names()
            },
            "total_rows": int(len(df)),
            "column_names": df.columns.tolist(),
            "column_types": self._get_column_types(df),
            "sample_data": sample_data,
            "column_stats": self._get_column_stats(df)
        }

    def read(
        self,
        sheet_name: Optional[str] = None,
        usecols: Optional[Union[str, List[int], List[str]]] = None
    ) -> Dict[str, Any]:
        """
        读取完整数据

        Args:
            sheet_name: 工作表名称（可选）
            usecols: 要读取的列（可选）

        Returns:
            包含完整数据的字典
        """
        df = self._read_file(sheet_name, usecols=usecols)
        return {
            "success": True,
            "data": self._df_to_array(df),
            "rows": int(len(df)),
            "columns": int(len(df.columns))
        }

    def query(
        self,
        filters: List[Dict] = None,
        group_by: str = None,
        aggregation: str = None,
        aggregate_column: str = None,
        order_by: str = None,
        order: str = "asc",
        limit: int = 100,
        sheet_name: str = None,
        usecols: Optional[Union[str, List[int], List[str]]] = None
    ) -> Dict[str, Any]:
        """
        查询并聚合数据，只返回聚合结果（不返回原始数据）

        重要：此方法设计用于 MCP 服务器端数据处理，
             只向大模型返回聚合后的摘要数据，不传输大量原始数据。

        Args:
            filters: 过滤条件列表
            group_by: 分组列名（必需）
            aggregation: 聚合函数 (sum, avg, count, min, max)（必需）
            aggregate_column: 要聚合的列名（count 时可选）
            order_by: 排序列名
            order: 排序方向 (asc, desc)
            limit: 返回的最大分组数（默认100）
            sheet_name: 工作表名称
            usecols: 要读取的列（可选）

        Returns:
            聚合结果字典（只包含分组和聚合值，不包含原始数据）

        Raises:
            ValueError: 如果未提供 group_by 和 aggregation
        """
        # 强制要求分组聚合，不返回原始数据
        if not group_by or not aggregation:
            logger.error("query 方法缺少必需参数: group_by 或 aggregation")
            raise ValueError(
                "query 方法必须指定 group_by 和 aggregation 参数。"
                "如需查看原始数据，请使用 read_head() 或 read_tail() 方法。"
            )

        logger.info(f"执行查询: group_by={group_by}, aggregation={aggregation}, filters={len(filters) if filters else 0}个")

        # 读取数据（只读取需要的列）
        df = self._read_file(sheet_name, usecols=usecols)

        # 应用过滤
        if filters:
            df = self._apply_filters(df, filters)

        # 分组聚合
        df = self._group_and_aggregate(df, group_by, aggregation, aggregate_column)

        # 排序
        if order_by and order_by in df.columns:
            df = df.sort_values(by=order_by, ascending=(order == "asc"))

        # 限制返回的分组数量
        df = df.head(limit)

        # 返回紧凑格式：标签和值分离
        if len(df.columns) == 2:
            # 标准聚合格式：[分组列, 聚合值列]
            labels = df.iloc[:, 0].tolist()
            values = df.iloc[:, 1].tolist()
            return {
                "success": True,
                "label": labels[0] if labels else None,
                "labels": labels,
                "values": values,
                "data": list(zip(labels, values)),
                "rows": int(len(df)),
                "aggregation": aggregation,
                "grouped_by": group_by
            }
        else:
            # 其他格式：返回完整数据
            return {
                "success": True,
                "data": self._df_to_array(df),
                "rows": int(len(df)),
                "aggregation": aggregation,
                "grouped_by": group_by
            }

    def read_chunked(
        self,
        chunksize: int = 10000,
        sheet_name: Optional[str] = None,
        usecols: Optional[Union[str, List[int], List[str]]] = None
    ):
        """
        分块读取大文件

        Args:
            chunksize: 每块的行数（默认10000）
            sheet_name: 工作表名称（可选）
            usecols: 要读取的列（可选）

        Returns:
            迭代器，每次生成一个 DataFrame 块

        Example:
            >>> reader = ExcelReader("large_file.xlsx")
            >>> for chunk in reader.read_chunked(chunksize=5000):
            ...     # 处理每个数据块
            ...     print(f"处理 {len(chunk)} 行数据")
        """
        suffix = self.file_path.suffix.lower()
        if suffix in ['.xlsx', '.xls']:
            return pd.read_excel(
                self.file_path,
                sheet_name=sheet_name or 0,
                engine='openpyxl',
                chunksize=chunksize,
                usecols=usecols
            )
        elif suffix == '.csv':
            return pd.read_csv(
                self.file_path,
                chunksize=chunksize,
                usecols=usecols
            )
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")

    def read_parallel_chunks(
        self,
        chunksize: int = 10000,
        num_chunks: int = 4,
        sheet_name: Optional[str] = None,
        usecols: Optional[Union[str, List[int], List[str]]] = None
    ) -> List[pd.DataFrame]:
        """
        多线程并行读取多个数据块

        Args:
            chunksize: 每块的行数（默认10000）
            num_chunks: 要读取的块数（默认4）
            sheet_name: 工作表名称（可选）
            usecols: 要读取的列（可选）

        Returns:
            包含多个 DataFrame 块的列表

        Example:
            >>> reader = ExcelReader("large_file.xlsx", enable_threading=True)
            >>> chunks = reader.read_parallel_chunks(chunksize=5000, num_chunks=4)
            >>> df_combined = pd.concat(chunks, ignore_index=True)
        """
        if not self.enable_threading:
            raise ValueError("多线程读取未启用，请在初始化时设置 enable_threading=True")

        from concurrent.futures import ThreadPoolExecutor

        def _read_chunk(skip_rows: int) -> pd.DataFrame:
            """读取单个数据块的内部函数"""
            suffix = self.file_path.suffix.lower()
            if suffix in ['.xlsx', '.xls']:
                return pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name or 0,
                    engine='openpyxl',
                    skiprows=range(1, skip_rows + 1) if skip_rows > 0 else None,
                    nrows=chunksize,
                    usecols=usecols
                )
            elif suffix == '.csv':
                return pd.read_csv(
                    self.file_path,
                    skiprows=range(1, skip_rows + 1) if skip_rows > 0 else None,
                    nrows=chunksize,
                    usecols=usecols
                )
            else:
                raise ValueError(f"不支持的文件格式: {suffix}")

        # 使用线程池并行读取
        with ThreadPoolExecutor(max_workers=min(num_chunks, 8)) as executor:
            skip_rows_list = [i * chunksize for i in range(num_chunks)]
            chunks = list(executor.map(_read_chunk, skip_rows_list))

        return chunks

    def get_row_count(
        self,
        sheet_name: Optional[str] = None
    ) -> int:
        """
        快速获取文件行数（不加载全部数据）

        Args:
            sheet_name: 工作表名称（可选）

        Returns:
            行数
        """
        cache_key = (sheet_name or "default",)

        # 检查缓存
        if self.enable_cache:
            mtime = self.file_path.stat().st_mtime
            if cache_key in self._row_count_cache:
                cached_count, cached_mtime = self._row_count_cache[cache_key]
                if cached_mtime == mtime:
                    return cached_count

        suffix = self.file_path.suffix.lower()
        if suffix in ['.xlsx', '.xls']:
            # Excel 文件使用 openpyxl 只读取行数
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            sheet = wb[sheet_name or wb.sheetnames[0]]
            count = sheet.max_row - 1  # 减去表头
            wb.close()
        elif suffix == '.csv':
            # CSV 文件使用 pandas 只读取行数
            df = pd.read_csv(self.file_path, nrows=0)
            # 使用更高效的方式统计行数
            with open(self.file_path, 'r', encoding='utf-8') as f:
                count = sum(1 for _ in f) - 1  # 减去表头
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")

        # 更新缓存
        if self.enable_cache:
            self._row_count_cache[cache_key] = (count, mtime)

        return count

    def read_head(
        self,
        n: int = 10,
        sheet_name: Optional[str] = None,
        usecols: Optional[Union[str, List[int], List[str]]] = None
    ) -> Dict[str, Any]:
        """
        快速读取文件前 n 行

        Args:
            n: 读取行数（默认10）
            sheet_name: 工作表名称（可选）
            usecols: 要读取的列（可选）

        Returns:
            包含前 n 行数据的字典
        """
        suffix = self.file_path.suffix.lower()
        if suffix in ['.xlsx', '.xls']:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name or 0,
                engine='openpyxl',
                nrows=n,
                usecols=usecols
            )
        elif suffix == '.csv':
            df = pd.read_csv(self.file_path, nrows=n, usecols=usecols)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")

        return {
            "success": True,
            "data": self._df_to_array(df),
            "rows": int(len(df)),
            "columns": int(len(df.columns))
        }

    def read_tail(
        self,
        n: int = 10,
        sheet_name: Optional[str] = None,
        usecols: Optional[Union[str, List[int], List[str]]] = None
    ) -> Dict[str, Any]:
        """
        快速读取文件后 n 行

        Args:
            n: 读取行数（默认10）
            sheet_name: 工作表名称（可选）
            usecols: 要读取的列（可选）

        Returns:
            包含后 n 行数据的字典
        """
        suffix = self.file_path.suffix.lower()
        if suffix in ['.xlsx', '.xls']:
            # Excel 使用 skipfooter 读取尾部
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name or 0,
                engine='openpyxl',
                usecols=usecols
            )
            df = df.tail(n)
        elif suffix == '.csv':
            # CSV 使用 skiprows 读取尾部
            total_rows = sum(1 for _ in open(self.file_path, 'r', encoding='utf-8'))
            skip_rows = max(0, total_rows - n - 1)  # 减去表头
            if skip_rows > 0:
                df = pd.read_csv(self.file_path, skiprows=range(1, skip_rows + 1), usecols=usecols)
            else:
                df = pd.read_csv(self.file_path, usecols=usecols)
            df = df.tail(n)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")

        return {
            "success": True,
            "data": self._df_to_array(df),
            "rows": int(len(df)),
            "columns": int(len(df.columns))
        }

    # ========== 私有方法 ==========

    @staticmethod
    def _parse_usecols(usecols: Optional[Union[str, List[int], List[str]]]) -> Optional[Union[List[int], List[str]]]:
        """
        解析 usecols 参数，支持逗号分隔的字符串

        Args:
            usecols: 可以是：
                - None: 读取所有列
                - 列字母范围字符串: "A:C"（pandas 原生支持）
                - 列索引列表: [0, 2, 3]
                - 列名列表: ["Col1", "Col2"]
                - 逗号分隔的列名字符串: "Col1,Col2"（本方法支持）

        Returns:
            解析后的 usecols，None 表示读取所有列
        """
        if usecols is None:
            return None

        # 如果已经是列表，直接返回
        if isinstance(usecols, list):
            return usecols

        # 如果是字符串，尝试解析
        if isinstance(usecols, str):
            # 检查是否是列字母范围（如 "A:C"），pandas 原生支持
            if ':' in usecols and usecols.replace(':', '').replace(' ', '').isalpha():
                return usecols

            # 否则按逗号分隔成列名列表
            # 去除空格并分割
            cols = [col.strip() for col in usecols.split(',')]
            # 过滤空字符串
            cols = [col for col in cols if col]
            return cols if cols else None

        # 其他类型（如单个整数）直接返回
        return usecols

    def _normalize_cache_key(self, sheet_name: str, usecols) -> str:
        """
        生成标准化的缓存键，列名排序确保一致性

        解决问题：['A','B'] 和 ['B','A'] 应该使用相同的缓存键

        Args:
            sheet_name: 工作表名称
            usecols: 列选择参数（解析后）

        Returns:
            标准化的缓存键字符串
        """
        sheet = sheet_name or 'default'
        if usecols is None:
            return f"{sheet}:full"

        # 对于字符串列名列表，排序后生成键
        if isinstance(usecols, list) and usecols and isinstance(usecols[0], str):
            sorted_cols = sorted(usecols)
            return f"{sheet}:cols:{','.join(sorted_cols)}"

        # 对于索引列表或其他格式，保持原有逻辑
        return f"{sheet}:cols:{str(usecols)}"

    def _read_file(
        self,
        sheet_name: Optional[str] = None,
        usecols: Optional[Union[str, List[int], List[str]]] = None,
        nrows: Optional[int] = None
    ) -> pd.DataFrame:
        """读取文件（支持缓存和 LRU 清理）"""
        # 解析 usecols 参数（支持逗号分隔的字符串）
        parsed_usecols = self._parse_usecols(usecols)

        # 使用标准化的缓存键
        cache_key = self._normalize_cache_key(sheet_name, parsed_usecols)

        # 检查缓存
        if self.enable_cache:
            mtime = self.file_path.stat().st_mtime
            # 1. 首先检查精确匹配的缓存
            if cache_key in self._cache:
                cached_df, cached_mtime = self._cache[cache_key]
                if cached_mtime == mtime:
                    # 缓存命中，移动到末尾（最近使用）
                    self._cache.move_to_end(cache_key)
                    self._cache_hits += 1

                    # 估算缓存数据大小
                    bytes_read = cached_df.memory_usage(deep=True).sum()
                    metrics.log_file_read(str(self.file_path), bytes_read, from_cache=True)

                    logger.debug(f"缓存命中: {cache_key} ({self._format_bytes(bytes_read)})")
                    return cached_df.copy()  # 返回副本避免修改缓存

            # 2. 智能缓存共享：当请求列子集时，检查全量缓存是否存在
            if parsed_usecols is not None:
                full_key = self._normalize_cache_key(sheet_name, None)
                if full_key in self._cache:
                    full_df, full_mtime = self._cache[full_key]
                    if full_mtime == mtime:
                        # 从全量缓存提取列子集
                        try:
                            subset_df = full_df[parsed_usecols].copy()
                            # 将子集缓存起来，下次直接命中
                            self._cache[cache_key] = (subset_df.copy(), mtime)
                            self._cache.move_to_end(cache_key)
                            self._cache_hits += 1

                            bytes_read = subset_df.memory_usage(deep=True).sum()
                            metrics.log_file_read(str(self.file_path), bytes_read, from_cache=True)

                            logger.debug(f"缓存共享命中: {cache_key} <- {full_key} ({self._format_bytes(bytes_read)})")
                            return subset_df.copy()
                        except KeyError:
                            # 请求的列不在全量缓存中，继续读取文件
                            logger.debug(f"缓存共享失败: 请求的列不在全量缓存中")

            self._cache_misses += 1
            logger.debug(f"缓存未命中: {cache_key}")

        suffix = self.file_path.suffix.lower()
        logger.debug(f"读取文件: {self.file_path.name}, 格式={suffix}, sheet={sheet_name or 'default'}")
        if suffix in ['.xlsx', '.xls']:
            # 明确指定 sheet_name=0 确保返回单个 DataFrame
            # 注意：必须使用 parsed_usecols（已解析为列表），而不是原始的 usecols 字符串
            # pandas read_excel 不接受逗号分隔的字符串作为 usecols 参数
            result = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name or 0,
                engine='openpyxl',
                usecols=parsed_usecols,
                nrows=nrows
            )
            # 如果返回的是字典（多sheet情况），取第一个
            if isinstance(result, dict):
                df = list(result.values())[0]
            else:
                df = result
        elif suffix == '.csv':
            # 注意：必须使用 parsed_usecols（已解析为列表），而不是原始的 usecols 字符串
            df = pd.read_csv(self.file_path, usecols=parsed_usecols)
        else:
            logger.error(f"不支持的文件格式: {suffix}")
            raise ValueError(f"不支持的文件格式: {suffix}")

        # 记录文件读取指标
        bytes_read = df.memory_usage(deep=True).sum()
        metrics.log_file_read(str(self.file_path), bytes_read, from_cache=False)
        logger.debug(f"文件读取完成: {self._format_bytes(bytes_read)}")

        # 更新缓存（使用 LRU 策略）
        if self.enable_cache:
            self._cache[cache_key] = (df.copy(), mtime)
            self._cache.move_to_end(cache_key)  # 新数据放到末尾
            # 超过缓存大小时，删除最旧的条目
            while len(self._cache) > self.max_cache_size:
                self._cache.popitem(last=False)

        return df

    @staticmethod
    def _format_bytes(bytes_size: int) -> str:
        """格式化字节大小"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if bytes_size < 1024.0:
                return f"{bytes_size:.1f}{unit}"
            bytes_size /= 1024.0
        return f"{bytes_size:.1f}TB"

    def _df_to_array(self, df: pd.DataFrame) -> List[List]:
        """将 DataFrame 转换为二维数组（包含表头），确保所有值 JSON 可序列化"""
        import numpy as np

        def _to_serializable(val):
            """将 numpy 类型转换为 JSON 可序列化的 Python 原生类型"""
            if pd.isna(val):
                return None
            elif isinstance(val, (np.integer, np.int64, np.int32)):
                return int(val)
            elif isinstance(val, (np.floating, np.float64, np.float32)):
                return float(val)
            elif isinstance(val, (np.bool_, bool)):
                return bool(val)
            elif hasattr(val, 'isoformat'):  # datetime
                return val.isoformat()
            else:
                return val

        # 复制 DataFrame 以避免修改原始数据
        df_copy = df.copy()

        # 将 datetime 类型转换为字符串
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                df_copy[col] = df_copy[col].astype(str)

        # 转换为列表并确保所有值都是 JSON 可序列化的
        result = [df_copy.columns.tolist()]
        for row in df_copy.values.tolist():
            result.append([_to_serializable(val) for val in row])

        return result

    def _get_column_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """获取列的数据类型"""
        type_map = {
            'object': 'string',
            'int64': 'number',
            'float64': 'number',
            'int32': 'number',
            'float32': 'number',
            'datetime64[ns]': 'datetime',
            'bool': 'boolean'
        }
        return {col: type_map.get(str(dtype), str(dtype)) for col, dtype in df.dtypes.items()}

    def _get_sheet_names(self) -> List[str]:
        """获取所有工作表名称（优化版，只读取元数据）"""
        if self.file_path.suffix.lower() in ['.xlsx', '.xls']:
            try:
                # 使用 openpyxl 只读取 sheet 名称，不解析内容
                import warnings
                from openpyxl import load_workbook
                # 抑制 openpyxl 的警告
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    wb = load_workbook(self.file_path, read_only=True, keep_links=False)
                    sheet_names = wb.sheetnames
                    wb.close()
                return sheet_names
            except Exception as e:
                # 如果 openpyxl 失败，回退到 pandas
                return pd.ExcelFile(self.file_path).sheet_names
        return ["Sheet1"]

    def _get_column_stats(self, df: pd.DataFrame) -> Dict[str, Any]:
        """获取每列的基本统计信息（优化版，只使用小样本）"""
        stats = {}
        # 只使用前 100 行计算统计信息，足够了解数据分布
        df_len = int(len(df))
        sample_df = df.head(min(100, df_len)) if df_len > 100 else df

        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                stats[col] = {
                    "min": float(df[col].min()) if not df[col].isna().all() else None,
                    "max": float(df[col].max()) if not df[col].isna().all() else None,
                    "mean": float(df[col].mean()) if not df[col].isna().all() else None
                }
            else:
                # 使用样本数据计算唯一值和示例值
                unique_vals = sample_df[col].dropna().unique()
                # 将 datetime 类型的值转换为字符串
                values_list = unique_vals[:10].tolist()
                values_list = [str(v) if not isinstance(v, (str, int, float, bool, type(None))) else v for v in values_list]
                # 唯一值数量也使用样本计算
                stats[col] = {
                    "unique_values": int(sample_df[col].nunique()),
                    "values": values_list
                }
        return stats

    def _apply_filters(self, df: pd.DataFrame, filters: List[Dict]) -> pd.DataFrame:
        """应用过滤条件"""
        for f in filters:
            col = f.get("column")
            op = f.get("operator", "=")
            val = f.get("value")

            if col not in df.columns:
                continue

            if op == "=":
                df = df[df[col] == val]
            elif op == "!=":
                df = df[df[col] != val]
            elif op == ">":
                df = df[df[col] > val]
            elif op == "<":
                df = df[df[col] < val]
            elif op == ">=":
                df = df[df[col] >= val]
            elif op == "<=":
                df = df[df[col] <= val]
            elif op == "contains":
                df = df[df[col].astype(str).str.contains(str(val), na=False)]
        return df

    def _group_and_aggregate(
        self,
        df: pd.DataFrame,
        group_by: str,
        aggregation: str,
        aggregate_column: str
    ) -> pd.DataFrame:
        """
        分组聚合

        支持的聚合函数:
        - 基础统计: sum, avg, count, min, max
        - 高级统计: median, std, var
        - 位置值: first, last
        - 基数统计: nunique
        """
        agg_funcs = {
            "sum": "sum",
            "avg": "mean",
            "count": "count",
            "min": "min",
            "max": "max",
            "median": "median",
            "std": "std",
            "var": "var",
            "first": "first",
            "last": "last",
            "nunique": "nunique"
        }

        if aggregation not in agg_funcs:
            raise ValueError(
                f"不支持的聚合函数: {aggregation}。"
                f"支持的函数: {', '.join(agg_funcs.keys())}"
            )

        agg_func = agg_funcs[aggregation]

        # count 和 nunique 不需要指定聚合列
        if aggregation in ("count", "nunique"):
            result = df.groupby(group_by).size().reset_index()
            result.columns = [group_by, aggregate_column or aggregation]
        else:
            if not aggregate_column:
                raise ValueError(f"聚合函数 '{aggregation}' 需要指定 aggregate_column 参数")
            result = df.groupby(group_by)[aggregate_column].agg(agg_func).reset_index()
            result.columns = [group_by, aggregate_column]

        return result

    def clear_cache(self) -> Dict[str, Any]:
        """
        清空所有缓存

        Returns:
            操作结果字典
        """
        cleared_count = len(self._cache) + len(self._row_count_cache)
        self._cache.clear()
        self._row_count_cache.clear()
        self._cache_hits = 0
        self._cache_misses = 0

        return {
            "success": True,
            "message": f"已清空 {cleared_count} 个缓存条目"
        }

    def get_cache_info(self) -> Dict[str, Any]:
        """
        获取缓存统计信息

        Returns:
            缓存统计信息字典
        """
        # 估算缓存内存大小
        cache_memory = 0
        for df, _ in self._cache.values():
            cache_memory += df.memory_usage(deep=True).sum()

        total_requests = self._cache_hits + self._cache_misses
        hit_rate = self._cache_hits / total_requests if total_requests > 0 else 0

        return {
            "success": True,
            "data_cache_count": len(self._cache),
            "row_count_cache_count": len(self._row_count_cache),
            "total_cache_entries": len(self._cache) + len(self._row_count_cache),
            "max_cache_size": self.max_cache_size,
            "cache_memory_bytes": cache_memory,
            "cache_memory_mb": round(cache_memory / 1024 / 1024, 2),
            "cache_hits": self._cache_hits,
            "cache_misses": self._cache_misses,
            "total_requests": total_requests,
            "hit_rate": round(hit_rate * 100, 2)
        }
